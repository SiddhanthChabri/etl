import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from db_connection import get_db_connection
from logger_config import setup_logger
from datetime import datetime
from data_quality_checks import validate_fact_sales

# Setup logger
logger = setup_logger('FactSales')

def load_fact_sales():
    """Load sales data from database to Excel with logging"""
    
    start_time = datetime.now()
    logger.info("="*60)
    logger.info("Starting Fact Sales Load Process")
    logger.info(f"Start Time: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("="*60)
    
    connection = None
    cursor = None
    
    try:
        # Step 1: Connect to database
        logger.info("Step 1: Connecting to database...")
        connection = get_db_connection()
        cursor = connection.cursor()
        logger.info("✓ Database connection established successfully")
        
        # Step 2: Execute query
        logger.info("Step 2: Executing query to fetch sales data...")
        query = """
        SELECT 
            sales_key,
            customer_key,
            product_key,
            store_key,
            time_key,
            quantity_sold,
            sales_amount,
            discount_amount,
            created_at
        FROM fact_sales
        ORDER BY sales_key
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        logger.info(f"✓ Query executed successfully")
        logger.info(f"✓ Total records fetched: {len(results)}")
        
        # Step 3: Create DataFrame
        logger.info("Step 3: Creating DataFrame from query results...")
        columns = [
            'sales_key', 
            'customer_key', 
            'product_key', 
            'store_key',
            'time_key', 
            'quantity_sold', 
            'sales_amount', 
            'discount_amount', 
            'created_at'
        ]
        df = pd.DataFrame(results, columns=columns)
        logger.info(f"✓ DataFrame created successfully")
        logger.info(f"✓ DataFrame shape: {df.shape[0]} rows x {df.shape[1]} columns")
        
        # NEW: Run data quality checks
        logger.info("\n" + "="*60)
        logger.info("Running Data Quality Checks...")
        logger.info("="*60)
        
        quality_passed = validate_fact_sales(df)
        
        if not quality_passed:
            logger.warning("Data quality issues detected! Review logs before proceeding.")
            # Uncomment below to stop on quality issues
            # raise Exception("Data quality validation failed")
        
        # Continue with existing code...
        logger.info("Step 4: Data summary:")
        logger.info(f"✓ DataFrame shape: {df.shape[0]} rows x {df.shape[1]} columns")
        
        # Step 4: Basic data info
        logger.info("Step 4: Data summary:")
        logger.info(f"   - Total Sales Amount: ₹{df['sales_amount'].sum():,.2f}")
        logger.info(f"   - Total Quantity Sold: {df['quantity_sold'].sum():,}")
        logger.info(f"   - Total Discount: ₹{df['discount_amount'].sum():,.2f}")
        logger.info(f"   - Date Range: {df['time_key'].min()} to {df['time_key'].max()}")
        
        # Step 5: Load Excel workbook
        logger.info("Step 5: Loading Excel workbook...")
        excel_filename = 'Retail_DW_Data.xlsx'
        
        try:
            wb = openpyxl.load_workbook(excel_filename)
            logger.info(f"✓ Excel workbook '{excel_filename}' loaded successfully")
        except FileNotFoundError:
            logger.warning(f"✗ Workbook '{excel_filename}' not found. Creating new workbook...")
            wb = openpyxl.Workbook()
            wb.save(excel_filename)
            logger.info(f"✓ New workbook '{excel_filename}' created")
        
        # Step 6: Handle existing sheet
        sheet_name = 'fact_sales'
        if sheet_name in wb.sheetnames:
            logger.info(f"Sheet '{sheet_name}' already exists - removing old data")
            del wb[sheet_name]
            logger.info(f"✓ Old '{sheet_name}' sheet removed")
        
        # Step 7: Create new sheet
        logger.info("Step 6: Creating new fact_sales sheet...")
        ws = wb.create_sheet(sheet_name)
        logger.info(f"✓ New '{sheet_name}' sheet created")
        
        # Step 8: Write data to Excel
        logger.info("Step 7: Writing data to Excel sheet...")
        row_count = 0
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
            row_count = r_idx
        
        logger.info(f"✓ Data written successfully - {row_count} rows (including header)")
        
        # Step 9: Save workbook
        logger.info("Step 8: Saving Excel workbook...")
        wb.save(excel_filename)
        logger.info(f"✓ Workbook '{excel_filename}' saved successfully")
        
        # Step 10: Close connections
        logger.info("Step 9: Closing database connections...")
        if cursor:
            cursor.close()
        if connection:
            connection.close()
        logger.info("✓ Database connections closed successfully")
        
        # Calculate duration
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        # Final summary
        logger.info("="*60)
        logger.info("✓✓✓ FACT SALES LOAD COMPLETED SUCCESSFULLY! ✓✓✓")
        logger.info("="*60)
        logger.info(f"Summary:")
        logger.info(f"   - Records Loaded: {len(df):,}")
        logger.info(f"   - Total Sales: ₹{df['sales_amount'].sum():,.2f}")
        logger.info(f"   - Duration: {duration:.2f} seconds")
        logger.info(f"   - End Time: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("="*60)
        
        return True
        
    except Exception as e:
        logger.error("="*60)
        logger.error("✗✗✗ ERROR IN FACT SALES LOAD PROCESS ✗✗✗")
        logger.error("="*60)
        logger.error(f"Error Type: {type(e).__name__}")
        logger.error(f"Error Message: {str(e)}")
        logger.error("="*60)
        
        # Close connections even on error
        if cursor:
            cursor.close()
        if connection:
            connection.close()
        
        raise
    
if __name__ == "__main__":
    try:
        load_fact_sales()
    except Exception as e:
        logger.error("Script execution failed. Check logs for details.")
        exit(1)
